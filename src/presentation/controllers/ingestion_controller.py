from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
import csv
import hashlib
import openpyxl
import io
import re
import unicodedata
from datetime import datetime, date, timedelta
from sqlalchemy.orm import Session
from src.infrastructure.database.config import get_db
from src.infrastructure.database import models
from src.application.services.ingestion_service import IngestionService
from src.application.dto.ingestion_schema import (
    AtendimentoTransacionalImportSchema,
    VoalleAgregadoImportSchema
)

router = APIRouter(prefix="/ingestion", tags=["Ingestão"])

CHUNK_SIZE = 6000
TURNOS_VALIDOS = ["Madrugada", "Manhã", "Tarde", "Noite"]
SETOR_PREFIXO = "SAC"

# ==========================================
# LIMITES DE SEGURANÇA
# ==========================================

MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024
DATA_MINIMA = date(2020, 1, 1)
DATA_MAXIMA_OFFSET_DIAS = 1
TEMPO_MAXIMO_SEGUNDOS = 86400


# ==========================================
# SANITIZAÇÃO DE ERROS
# ==========================================

def sanitizar_erro(erro: Exception) -> str:
    """
    Nunca expõe SQL, stack traces ou dados internos ao usuário.
    Retorna sempre uma mensagem limpa em português.
    """
    msg = str(erro).lower()

    # Detecta erros de SQL / banco
    if any(kw in msg for kw in ("sqlalchemy", "psycopg", "duplicate key", "unique constraint",
                                  "violates", "relation", "column", "parameters truncated",
                                  "background on this error", "insert into")):
        return (
            "Ocorreu um erro ao salvar os dados no banco. "
            "É possível que alguns registros já existam. "
            "Tente novamente ou entre em contato com o administrador."
        )

    # Detecta erros de conexão
    if any(kw in msg for kw in ("connection", "timeout", "refused", "unreachable")):
        return "Erro de conexão com o banco de dados. Tente novamente em alguns instantes."

    # Erro genérico — limita tamanho e remove caracteres técnicos
    texto = str(erro)
    if len(texto) > 200:
        texto = texto[:200] + "..."

    return f"Erro interno de processamento. Tente novamente ou entre em contato com o administrador."


# ==========================================
# FUNÇÕES AUXILIARES
# ==========================================

def calcular_turno(dt: datetime) -> str:
    hora = dt.hour
    if 0 <= hora <= 5:
        return "Madrugada"
    elif 6 <= hora <= 11:
        return "Manhã"
    elif 12 <= hora <= 17:
        return "Tarde"
    else:
        return "Noite"

def detect_encoding(raw: bytes) -> str:
    sample = raw[3:] if raw.startswith(b"\xef\xbb\xbf") else raw
    try:
        sample.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        pass
    try:
        import chardet # type: ignore
        result = chardet.detect(raw)
        enc = (result.get("encoding") or "utf-8").lower().replace("-", "_")
        return "utf-8" if "ascii" in enc else enc
    except Exception:
        return "utf-8"

def detect_separator(text: str) -> str:
    first_line = text.split("\n")[0]
    candidates = {",": first_line.count(","), ";": first_line.count(";")}
    max_count = max(candidates.values())
    return max(candidates, key=lambda k: candidates[k]) if max_count > 0 else ";"

def parse_time_to_seconds(time_str: str) -> int:
    if not time_str or time_str.strip() in ("-", ""):
        return 0
    try:
        parts = time_str.strip().split(":")
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        elif len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
        return 0
    except Exception:
        return 0

def normalizar_nome(nome: str) -> str:
    nfkd = unicodedata.normalize("NFKD", nome)
    sem_acento = "".join(c for c in nfkd if not unicodedata.combining(c))
    return " ".join(sem_acento.upper().split())

def clean_agent_name(name: str) -> str:
    if not name:
        return "Desconhecido"
    nome_sem_ramal = name.split(" - ")[0].strip()
    return normalizar_nome(nome_sem_ramal)

def safe_int(value) -> int:
    if not value or str(value).strip() == "-":
        return 0
    try:
        return int(float(str(value).replace(",", ".").strip()))
    except ValueError:
        return 0

def safe_float_or_none(value) -> float | None:
    if not value or str(value).strip() == "-":
        return None
    try:
        val = float(str(value).replace(",", ".").strip())
        return max(-999.99, min(999.99, val))
    except ValueError:
        return None

def extract_date_from_filename(filename: str) -> date:
    match = re.search(r"(\d{2,4}[-_]?\d{2}[-_]?\d{2,4})", filename)
    if match:
        date_str = match.group(1).replace("_", "").replace("-", "")
        try:
            if len(date_str) == 8:
                if date_str.startswith("20"):
                    return datetime.strptime(date_str, "%Y%m%d").date()
                else:
                    return datetime.strptime(date_str, "%d%m%Y").date()
        except Exception:
            pass
    return date.today()

def calcular_hash_arquivo(raw_content: bytes) -> str:
    return hashlib.sha256(raw_content).hexdigest()


# ==========================================
# VALIDAÇÕES
# ==========================================

def validar_data(dt: date | datetime) -> str | None:
    data_apenas = dt.date() if isinstance(dt, datetime) else dt
    data_maxima = date.today() + timedelta(days=DATA_MAXIMA_OFFSET_DIAS)

    if data_apenas < DATA_MINIMA:
        return (
            f"A data {data_apenas.strftime('%d/%m/%Y')} é anterior a "
            f"{DATA_MINIMA.strftime('%d/%m/%Y')}. "
            f"Por favor, verifique se o arquivo contém datas válidas."
        )
    if data_apenas > data_maxima:
        return (
            f"A data {data_apenas.strftime('%d/%m/%Y')} é uma data no futuro. "
            f"Por favor, insira uma data válida."
        )
    return None

def validar_tempo(segundos: int, tipo: str) -> str | None:
    if segundos > TEMPO_MAXIMO_SEGUNDOS:
        horas = segundos / 3600
        return (
            f"O tempo de {tipo} encontrado foi de {horas:.1f} horas, "
            f"o que excede o limite de 24 horas."
        )
    return None

def is_setor_permitido(valor: str) -> bool:
    if not valor or not valor.strip():
        return True
    return valor.strip().upper().startswith(SETOR_PREFIXO)

def detect_format(headers: list[str]):
    headers_upper = [h.upper() for h in headers]
    is_voalle_agregado = all(h in headers_upper for h in ["CA", "NA", "NSF"])
    is_omnichannel = "Nome do Atendente" in headers or "Tempo em Espera na Fila" in headers
    is_ligacao = "Data de início" in headers and "Sentido" in headers
    return is_voalle_agregado, is_omnichannel, is_ligacao


# ==========================================
# PRÉ-VALIDAÇÃO
# ==========================================

def pre_validar_planilha(rows, is_voalle, is_omnichannel, is_ligacao, voalle_data_ref=None):
    erros = []
    if is_voalle:
        return erros

    for index, row in enumerate(rows, start=1):
        if len(erros) >= 5:
            erros.append("... e possivelmente mais linhas com problemas semelhantes.")
            break
        try:
            if is_omnichannel:
                data_str_raw = (row.get("Data Inicial") or "").strip()
                hora_str_raw = (row.get("Hora Inicial") or "").strip()
                if not data_str_raw:
                    continue
                data_completa = f"{data_str_raw} {hora_str_raw}".strip()
                try:
                    dt = datetime.strptime(data_completa, "%d/%m/%Y %H:%M:%S")
                except ValueError:
                    erros.append(f"Linha {index}: A data '{data_completa}' não está no formato esperado (DD/MM/AAAA HH:MM:SS).")
                    continue
            elif is_ligacao:
                data_str_raw = (row.get("Data de início") or "").strip()
                if not data_str_raw:
                    continue
                try:
                    dt = datetime.strptime(data_str_raw, "%d/%m/%Y %H:%M:%S")
                except ValueError:
                    erros.append(f"Linha {index}: A data '{data_str_raw}' não está no formato esperado (DD/MM/AAAA HH:MM:SS).")
                    continue
            else:
                continue

            erro_data = validar_data(dt)
            if erro_data:
                erros.append(f"Linha {index}: {erro_data}")
                continue

            if is_omnichannel:
                te = parse_time_to_seconds(row.get("Tempo em Espera na Fila") or "")
                ta = parse_time_to_seconds(row.get("Tempo em Atendimento") or "")
            else:
                te = parse_time_to_seconds(row.get("Espera") or "")
                ta = parse_time_to_seconds(row.get("Atendimento") or "")

            erro_e = validar_tempo(te, "espera")
            if erro_e:
                erros.append(f"Linha {index}: {erro_e}")
            erro_a = validar_tempo(ta, "atendimento")
            if erro_a:
                erros.append(f"Linha {index}: {erro_a}")
        except Exception:
            pass

    return erros


# ==========================================
# PARSE DE LINHAS PARA DTO
# ==========================================

def parse_row_to_dto(row, is_voalle, is_omnichannel, is_ligacao, voalle_data_ref=None):
    if is_voalle:
        assert voalle_data_ref is not None
        return VoalleAgregadoImportSchema(
            data_referencia=voalle_data_ref,
            colaborador_nome=clean_agent_name((row.get("Atendente") or "Desconhecido").strip()),
            clientes_atendidos=safe_int(row.get("CA")),
            numero_atendimentos=safe_int(row.get("NA")),
            solicitacao_finalizada=safe_int(row.get("NSF")),
        )
    elif is_omnichannel:
        equipe_raw = (row.get("Nome da Equipe") or "").strip()
        if not is_setor_permitido(equipe_raw):
            return None
        data_inicial = (row.get("Data Inicial") or "").strip()
        hora_inicial = (row.get("Hora Inicial") or "").strip()
        data_str = f"{data_inicial} {hora_inicial}".strip()
        data_ref = datetime.strptime(data_str, "%d/%m/%Y %H:%M:%S") if data_str else datetime.now()
        return AtendimentoTransacionalImportSchema(
            data_referencia=data_ref,
            turno=calcular_turno(data_ref),
            colaborador_nome=clean_agent_name((row.get("Nome do Atendente") or "Desconhecido").strip()),
            equipe=(equipe_raw or None),
            canal_nome="WhatsApp",
            status_nome=(row.get("Status") or "Desconhecido").strip(),
            protocolo=(row.get("Número do Protocolo") or None),
            sentido_interacao=None,
            tempo_espera_segundos=parse_time_to_seconds(row.get("Tempo em Espera na Fila") or ""),
            tempo_atendimento_segundos=parse_time_to_seconds(row.get("Tempo em Atendimento") or ""),
            nota_solucao=safe_float_or_none(row.get("Avaliação - Nota da Solução Oferecida")),
            nota_atendimento=safe_float_or_none(row.get("Avaliação - Nota do Atendimento Prestado")),
        )
    elif is_ligacao:
        fila_raw = (row.get("Fila") or "").strip()
        if not is_setor_permitido(fila_raw):
            return None
        data_inicio = (row.get("Data de início") or "").strip()
        data_ref = datetime.strptime(data_inicio, "%d/%m/%Y %H:%M:%S")
        return AtendimentoTransacionalImportSchema(
            data_referencia=data_ref,
            turno=calcular_turno(data_ref),
            colaborador_nome=clean_agent_name((row.get("Agente") or "Desconhecido").strip()),
            equipe=(fila_raw or None),
            canal_nome="Ligação",
            status_nome=(row.get("Status") or "Desconhecido").strip(),
            protocolo=(row.get("Protocolo") or None),
            sentido_interacao=(row.get("Sentido") or None),
            tempo_espera_segundos=parse_time_to_seconds(row.get("Espera") or ""),
            tempo_atendimento_segundos=parse_time_to_seconds(row.get("Atendimento") or ""),
            nota_solucao=None,
            nota_atendimento=safe_float_or_none(row.get("Avaliação 1")),
        )
    raise ValueError("Formato não reconhecido")


# ==========================================
# ENDPOINT: UPLOAD DE PLANILHA
# ==========================================

@router.post("/upload-csv")
async def upload_csv(
    file: UploadFile = File(...),
    data_voalle: str = Form(None),
    db: Session = Depends(get_db)
):
    if not file.filename or not file.filename.endswith((".csv", ".xlsx")):
        raise HTTPException(
            status_code=400,
            detail="Formato de arquivo não permitido. Por favor, envie um arquivo .csv ou .xlsx."
        )

    service = IngestionService(db)

    # ── 1. Ler e validar tamanho ──
    raw_content = await file.read()

    if len(raw_content) == 0:
        raise HTTPException(
            status_code=400,
            detail="O arquivo enviado está vazio. Por favor, selecione um arquivo com dados."
        )

    if len(raw_content) > MAX_FILE_SIZE_BYTES:
        tamanho_mb = len(raw_content) / 1024 / 1024
        limite_mb = MAX_FILE_SIZE_BYTES / 1024 / 1024
        raise HTTPException(
            status_code=400,
            detail=f"O arquivo tem {tamanho_mb:.1f} MB e excede o limite de {limite_mb:.0f} MB. Divida em partes menores."
        )

    # ── 2. Hash + verificar duplicidade ──
    file_hash = calcular_hash_arquivo(raw_content)

    if service.verificar_hash_duplicado(file_hash):
        raise HTTPException(
            status_code=409,
            detail="Este arquivo já foi importado anteriormente. Envie um arquivo diferente para evitar duplicidade de dados."
        )

    # ── 3. Registrar upload (pending) ──
    upload_record = models.Upload(
        file_path=file.filename,
        file_hash=file_hash,
        status="pending",
        created_at=datetime.utcnow(),
    )
    db.add(upload_record)
    db.commit()

    try:
        rows = []
        headers = []

        if file.filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(raw_content), data_only=True, read_only=True)
            sheet = wb.active
            if sheet is None:
                raise HTTPException(status_code=400, detail="Planilha Excel vazia ou inválida.")
            raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip() for h in raw_headers if h is not None]
            for row_values in sheet.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row_values):
                    continue
                rows.append({
                    headers[i]: str(row_values[i]).strip() if row_values[i] is not None else ""
                    for i in range(len(headers))
                })
            wb.close()
        else:
            encoding = detect_encoding(raw_content)
            try:
                decoded_content = raw_content.decode(encoding)
            except Exception:
                decoded_content = raw_content.decode("iso-8859-1", errors="replace")
            if decoded_content.startswith("\ufeff"):
                decoded_content = decoded_content[1:]
            separator = detect_separator(decoded_content)
            reader = csv.DictReader(io.StringIO(decoded_content), delimiter=separator)
            headers = [str(h).strip() for h in (reader.fieldnames or [])]
            rows = [
                {k.strip(): (v.strip() if v else "") for k, v in row.items()}
                for row in reader
            ]

        if not rows:
            upload_record.status = "error"
            upload_record.error = "Arquivo sem dados"
            upload_record.processed_at = datetime.utcnow()
            db.commit()
            raise HTTPException(
                status_code=400,
                detail="O arquivo não contém registros de dados. Verifique a planilha e tente novamente."
            )

        is_voalle, is_omnichannel, is_ligacao = detect_format(headers)

        if not (is_voalle or is_omnichannel or is_ligacao):
            raise HTTPException(
                status_code=400,
                detail="Não foi possível identificar o tipo da planilha. Verifique se as colunas estão corretas."
            )

        formato_nome = "Voalle" if is_voalle else "Omnichannel" if is_omnichannel else "Ligação"

        # ── Validar data Voalle ──
        voalle_data_ref = None
        if is_voalle:
            if data_voalle:
                try:
                    voalle_data_ref = datetime.strptime(data_voalle, "%Y-%m-%d").date()
                except ValueError:
                    raise HTTPException(
                        status_code=400,
                        detail=f"A data '{data_voalle}' não é válida. Selecione uma data válida no calendário."
                    )
            else:
                voalle_data_ref = extract_date_from_filename(file.filename)

            if voalle_data_ref is None:
                raise HTTPException(
                    status_code=400,
                    detail="Para planilhas Voalle, preencha o campo 'Data do Relatório'."
                )

            erro_data = validar_data(voalle_data_ref)
            if erro_data:
                raise HTTPException(status_code=400, detail=erro_data)

        # ── Pré-validação de datas/tempos ──
        erros_validacao = pre_validar_planilha(rows, is_voalle, is_omnichannel, is_ligacao, voalle_data_ref)
        if erros_validacao:
            upload_record.status = "error"
            upload_record.error = "; ".join(erros_validacao[:3])
            upload_record.processed_at = datetime.utcnow()
            db.commit()
            raise HTTPException(
                status_code=422,
                detail={
                    "tipo": "validacao",
                    "mensagem": f"A planilha ({formato_nome}) contém dados inválidos. Corrija os problemas e tente novamente.",
                    "erros": erros_validacao,
                    "total_linhas": len(rows),
                }
            )

        # ── Pré-carregar dados existentes (dedup Python-level) ──
        protocolos_existentes_banco: set[str] = set()
        voalle_existentes: set[int] = set()

        if is_voalle and voalle_data_ref:
            voalle_existentes = service.carregar_voalle_existentes(voalle_data_ref)
        elif not is_voalle:
            protocolos_do_arquivo = []
            for row in rows:
                proto = (row.get("Número do Protocolo") or row.get("Protocolo") or "").strip()
                if proto:
                    protocolos_do_arquivo.append(proto)
            if protocolos_do_arquivo:
                protocolos_existentes_banco = service.carregar_protocolos_existentes(protocolos_do_arquivo)

        # ── Processar ──
        total_success = 0
        total_error = 0
        total_ignorados = 0
        total_duplicados = 0
        all_errors = []
        chunk = []
        protocolos_vistos: set[str] = set()

        def flush_chunk():
            nonlocal total_success, total_error, total_duplicados, all_errors, chunk
            if not chunk:
                return
            if is_voalle:
                res = service.process_voalle_batch(chunk, voalle_existentes=voalle_existentes)
            else:
                res = service.process_transacional_batch(chunk, protocolos_existentes_banco=protocolos_existentes_banco)
            total_success += res["success_count"]
            total_error += res["error_count"]
            total_duplicados += res.get("duplicate_count", 0)
            all_errors.extend(res.get("errors", []))
            chunk = []

        for index, row in enumerate(rows, start=1):
            try:
                dto = parse_row_to_dto(row, is_voalle, is_omnichannel, is_ligacao, voalle_data_ref)
                if dto is None:
                    total_ignorados += 1
                    continue

                protocolo_val = getattr(dto, "protocolo", None)
                if protocolo_val:
                    if protocolo_val in protocolos_vistos or protocolo_val in protocolos_existentes_banco:
                        total_duplicados += 1
                        continue
                    protocolos_vistos.add(protocolo_val)

                chunk.append(dto)
                if len(chunk) >= CHUNK_SIZE:
                    flush_chunk()
            except Exception as e:
                total_error += 1
                if len(all_errors) < 20:
                    all_errors.append(f"Linha {index}: {str(e)}")

        flush_chunk()

        # ── Resposta ──
        if total_success == 0 and total_duplicados > 0 and total_error == 0:
            status = "duplicate"
        elif total_success > 0 and total_error == 0:
            status = "success"
        elif total_success > 0:
            status = "warning"
        else:
            status = "error"

        partes_msg = []
        if total_success > 0:
            partes_msg.append(f"{total_success} registros importados com sucesso")
        if total_ignorados:
            partes_msg.append(f"{total_ignorados} ignorados (outros setores)")
        if total_duplicados:
            partes_msg.append(f"{total_duplicados} duplicados ignorados")
        if total_error:
            partes_msg.append(f"{total_error} erros encontrados")

        if status == "duplicate":
            mensagem = "Todos os registros desta planilha já existem no banco de dados. Nenhum dado novo foi importado."
        elif status == "error" and total_success == 0:
            mensagem = "Não foi possível importar nenhum registro. Verifique a planilha."
        else:
            mensagem = ". ".join(partes_msg) + "."

        upload_record.status = status
        upload_record.total_registros = total_success
        upload_record.total_duplicados = total_duplicados
        upload_record.processed_at = datetime.utcnow()
        if all_errors:
            upload_record.error = "; ".join(all_errors[:5])
        db.commit()

        return {
            "status": status,
            "message": mensagem,
            "detalhes": {
                "formato_detectado": formato_nome,
                "total_linhas_arquivo": len(rows),
                "success_count": total_success,
                "ignored_count": total_ignorados,
                "duplicate_count": total_duplicados,
                "error_count": total_error,
                "errors": all_errors[:10],
            }
        }

    except HTTPException:
        upload_record.status = "error"
        upload_record.processed_at = datetime.utcnow()
        db.commit()
        raise
    except Exception as e:
        # ═══════════════════════════════════════════════════
        # NUNCA expor SQL ou stack trace ao usuário
        # ═══════════════════════════════════════════════════
        upload_record.status = "error"
        upload_record.processed_at = datetime.utcnow()
        upload_record.error = str(e)[:500]  # log interno completo
        db.commit()
        raise HTTPException(
            status_code=500,
            detail=sanitizar_erro(e)  # mensagem limpa pro usuário
        )


# ==========================================
# ENDPOINT: LISTAR COLABORADORES
# ==========================================

@router.get("/colaboradores")
def listar_colaboradores(db: Session = Depends(get_db)):
    colaboradores = db.query(models.DimColaborador).order_by(
        models.DimColaborador.turno,
        models.DimColaborador.nome
    ).all()
    return [
        {"id": c.id, "nome": c.nome, "equipe": c.equipe, "turno": c.turno or "Não calculado"}
        for c in colaboradores
    ]